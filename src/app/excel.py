from openpyxl import load_workbook, Workbook
from copy import copy
import pandas as pd
import os
import re

class ExcelDataLoader:
    def __init__(self):
        self.file_path = None
        self.excel = None
        self.sheets = None
    
    def load_file(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # print(f"📁 Loading Excel file: {file_path}")
        self.file_path = file_path
        self.excel = pd.ExcelFile(file_path)
        self.sheets = self.excel.sheet_names
        # return self.excel

    def show_summary(self):
        print("\n===== 📊 FILE SUMMARY =====")
        print(f"File Name       : {os.path.basename(self.file_path)}")
        print(f"Number of Sheets: {len(self.sheets)}")
        print("Sheet Infor:")
        for i, sheet in enumerate(self.sheets, start=1):
            print(f"{i} - {sheet}")
        print("===========================\n")
    
    def load_sheet(self, sheet_number = None):
        # print(len(self.sheets) )

        if sheet_number < 1 or sheet_number > len(self.sheets):
            raise ValueError("Invalid sheet number.")

        df = pd.read_excel(self.file_path, sheet_name=self.sheets[sheet_number - 1])
        return df
    
    def load_template(self, sheet_number):
        print(len(self.sheets) )
        if sheet_number < 1 or sheet_number > len(self.sheets):
            raise ValueError("Invalid sheet number.")
        
        wb = load_workbook(self.file_path)
        ws = wb[self.sheets[sheet_number - 1]]
        return ws
    
    def show_info(self, df):
        print(f"\n===== 📄 SHEET INFO =====")
        print(f"Rows: {df.shape[0]} | Columns: {df.shape[1]}")

        print("\n🔹 Column Information:")
        print(df.dtypes.to_string())
        # print("\n🔹 Memory Usage:")
        # print(df.memory_usage(deep=True).sum(), "bytes")
        
        print("\n🔹 Xem trước:")
        print(df.head())
        print("==="*20)

        print("\nColumns:")
        for i, col in enumerate(df.columns, start=1):
            print(f'\t[{i}] "{col}"')



        
    def deep_inspect(self):
        """Show summary and detailed sheet info."""
        print('========= 📊 DEEP INSPECT =========')
        print(f"\n📁 Excel file: {file_path}")
        print(f"📄 Total sheets: {len(self.excel.sheet_names)}\n")

        total_rows = 0

        for idx, sheet_name in enumerate(self.excel.sheet_names, start=1):
            df = self.excel.parse(sheet_name)

            rows, cols = df.shape
            total_rows += rows

            col_names = ", ".join(df.columns.astype(str))

            print(
                f"{idx}) \"{sheet_name}\", "
                f"Size: {rows} - {cols}, "
                f"Columns: {col_names}"
            )

        print(f"\n📊 Total rows across all sheets: {total_rows}")


class DataProcessing:
    def __init__(self, base_name, raw_df, raw_ws, output_folder="data/output"):
        self.name = base_name
        self.raw_ws = raw_ws  # For template
        self.raw_df = raw_df
        self.col_list = self.raw_df[0].columns.tolist()
        self.output_name = "output_" + base_name
        self.output_folder = output_folder

    def _get_column_by_index(self, indexes):
        cols = []

        for i in indexes:
            if 1 <= i <= len(self.col_list):
                cols.append(self.col_list[i - 1])

        print("Pocessing with column: ", cols)
        return cols
    
    def delete_by_index(self, delete_indexes=None):
        """
        delete_indexes: list of column numbers to delete (1-based)
        sort_order_indexes: list of column numbers in the desired order (1-based)
        """
        # Convert 1-based indexes to column names
        # self.

        delete_cols = self._get_column_by_index(delete_indexes)
        # --- DELETE COLUMNS ---
        df = self.raw_df[0]
        df = df.drop(columns=delete_cols, errors='ignore')
        print(f"🗑 Deleted columns: {delete_indexes}")

        output = [("Sheet1", df)]
        self.save_to_file(output)
        # return df
    def reorder_by_index(self, sort_order_indexes=None):  
        # --- SORT COLUMNS ---
        if sort_order_indexes:
            # Map index → column name (only if column still exists)
            sorted_cols = []
            for i in sort_order_indexes:
                col_name = self._get_column_by_index(i)
                if col_name in self.col_list:
                    sorted_cols.append(col_name)

            # Add remaining columns at the end (not sorted)
            remaining = [c for c in self.col_list if c not in sorted_cols]

            df = self.raw_df[0][sorted_cols + remaining]
            output = [("Sheet1", df)]
            self.save_to_file(output)
            print(f"🔃 Sorted columns in order: {sort_order_indexes}")
            # return df
    
    def save_to_file(self, df_list, prefix="", cache=False):
        # Save temp rawdata
        output_cache = "temp.xlsx"
        with pd.ExcelWriter(output_cache, engine="openpyxl") as writer:
            for name, df in df_list:
                sheet_name = str(name)[:31]
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )
        print(f"Create temp file: {output_cache}")

        # Load rawdata output 
        wb_target = load_workbook(output_cache)

        # Apply template
        for sheet in wb_target.sheetnames:
            self.apply_style_from_template(
                wb_target, sheet
            )

        #Save to file
        file_name=os.path.join(self.output_folder,  f"{prefix}output_{self.name}")
        wb_target.save(file_name)
        print("Output file is saved to:  " + file_name)
        
        # Remove after processing

        if os.path.exists(output_cache) and not cache:
            os.remove(output_cache)
            print(f"Removed temp file: {output_cache}")

    def split_to_sheets_by_col(self, col_index):
        # Output file
        output_file = f"data/output_cache/temp.xlsx"

        category_col = self._get_column_by_index(col_index)[0]

        def extract_name(text):
            parts = text.split(" - ")
            if len(parts) < 2:
                return None

            s = parts[1]                     # "QTDND xã Thịnh Sơn"
            s = re.sub(r"^(QTDND\s+)?(xã|phường|thị trấn)\s+", "QTDND ", s, flags=re.IGNORECASE)
            return s.strip()
        

        def normalize_qtdnd_name(text: str) -> str:
            text = text.strip().upper()

            # 1. Remove fixed prefix
            prefix = "QUY TIN DUNG NHAN DAN"
            if not text.startswith(prefix):
                raise ValueError("Invalid QTDND name")

            s = text[len(prefix):].strip()

            # 2. Remove administrative units
            # Order matters: THI TRAN before TRAN
            s = re.sub(r"^(THI\s+TRAN|XA|PHUONG)\s+", "", s)

            # 3. Normalize spaces
            s = " ".join(s.split())

            return f"QTDND {s}"

        
        output_df = [
            (extract_name(category), group_df.reset_index(drop=True))
            for category, group_df in self.raw_df[0].groupby(category_col)
        ]
       

        # Preview 
        sheet_names = [name for name, _ in output_df[:5]]
        print(sheet_names)

        # Save output and apply format
        self.save_to_file(output_df)
    
    def merge_table(self):
        if not self.raw_df:
            raise ValueError("df_list is empty")

        # Optional safety check: ensure same columns
        base_cols = self.raw_df[0].columns
        for i, df in enumerate(self.raw_df, start=1):
            if not df.columns.equals(base_cols):
                raise ValueError(f"Schema mismatch in DataFrame #{i}")

        merged_df = pd.concat(self.raw_df, ignore_index=True)

        self.save_to_file([("sheet1", merged_df)])

    def apply_style_from_template(self, wb_target, sheet_target, header_row=1):
        """
        Apply formatting from template sheet to target sheet.

        Rules:
        1. Header row styles are copied exactly from template.
        2. Data rows use the style of the FIRST data row in template
        and apply it to ALL data rows in target.
        """

        ws_t = self.raw_ws
        ws_b = wb_target[sheet_target]

        # -------------------------
        # Sheet-level properties
        # -------------------------
        for col, dim in ws_t.column_dimensions.items():
            ws_b.column_dimensions[col].width = dim.width

        for row, dim in ws_t.row_dimensions.items():
            ws_b.row_dimensions[row].height = dim.height

        ws_b.freeze_panes = ws_t.freeze_panes

        for merged_range in ws_t.merged_cells.ranges:
            ws_b.merge_cells(str(merged_range))

        # Copy sheet view attributes individually
        # ws_b.sheet_view.zoomScale = ws_t.sheet_view.zoomScale
        # ws_b.sheet_view.workbookViewId = ws_t.sheet_view.workbookViewId
        # ws_b.sheet_view.showGridLines = ws_t.sheet_view.showGridLines
        # ws_b.sheet_view.rightToLeft = ws_t.sheet_view.rightToLeft

        # -------------------------
        # Determine column range
        # -------------------------
        max_col = min(ws_t.max_column, ws_b.max_column)

        # ==================================================
        # 1️⃣ Apply HEADER styles
        # ==================================================
        for col in range(1, max_col + 1):
            cell_t = ws_t.cell(row=header_row, column=col)
            cell_b = ws_b.cell(row=header_row, column=col)

            if not cell_t.has_style:
                continue

            cell_b.font = copy(cell_t.font)
            cell_b.border = copy(cell_t.border)
            cell_b.fill = copy(cell_t.fill)
            cell_b.alignment = copy(cell_t.alignment)
            cell_b.number_format = cell_t.number_format

        # ==================================================
        # 2️⃣ Apply DATA ROW styles (from FIRST data row)
        # ==================================================
        template_data_row = header_row + 1

        if template_data_row > ws_t.max_row:
            # Template has no data rows → nothing more to apply
            return

        for row in range(header_row + 1, ws_b.max_row + 1):
            for col in range(1, max_col + 1):
                cell_t = ws_t.cell(row=template_data_row, column=col)
                cell_b = ws_b.cell(row=row, column=col)

                if not cell_t.has_style:
                    continue

                cell_b.font = copy(cell_t.font)
                cell_b.border = copy(cell_t.border)
                cell_b.fill = copy(cell_t.fill)
                cell_b.alignment = copy(cell_t.alignment)
                cell_b.number_format = cell_t.number_format

            
    def inspect_all(self):
        """Show summary and detailed sheet info."""
        self.show_summary()

        for sheet in self.sheets:
            self.analyze_sheet(sheet)

def get_xlsx_files(folder_path: str):
    if not os.path.isdir(folder_path):
        raise ValueError("Invalid folder path")

    return [
        os.path.join(folder_path, f)
        for f in os.listdir(folder_path)
        if f.lower().endswith(".xlsx")
    ]

if __name__ == "__main__":
    file_path = "test2.xlsx"
    sheet_index = 1
    output_file = "test1_output.xlsx"

    # merge_file test
    folder = "data/input"
    excel_files = get_xlsx_files(folder)
    print(excel_files)
    loader = ExcelDataLoader()
    loader.load_file(excel_files[2])
    input_sheets = [loader.load_sheet()]
    ws_template = loader.load_template(sheet_index)

    for f in excel_files[1:]:
        print(f)
        loader.load_file(f)
        sheet = loader.load_sheet(sheet_index)
        input_sheets.append(sheet)
    
    print(len(input_sheets))
    data = DataProcessing("test1.xlsx", input_sheets, ws_template)
    data.merge_table()


    # loader = ExcelDataLoader(file_path)
    # loader.load_file()
    # loader.show_summary()
    # loader.deep_inspect()

    # sheet = loader.load_sheet(sheet_index)
    # ws_template = loader.load_template(sheet_index)
    # data.reorder_by_index([2, 1])

    # Delete test
    # loader = ExcelDataLoader()
    # loader.load_file("data/input/test1.xlsx")
    # sheet = loader.load_sheet(1)
    # ws_template = loader.load_template(sheet_index)
    # data = DataProcessing(file_path, [sheet], ws_template)
    # data.delete_by_index([1, 2])
   

    # Split test
    # loader = ExcelDataLoader()
    # loader.load_file("input.xlsx")
    # sheet = loader.load_sheet(1)
    # ws_template = loader.load_template(sheet_index)
    # data = DataProcessing(file_path, [sheet], ws_template)
    # data.split_to_sheets_by_col(3) 